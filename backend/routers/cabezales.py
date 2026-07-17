from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
import pandas as pd
import io
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from database import get_db
from models import CabezalModel, AlineacionCabezalModel, UserModel
from schemas import CabezalUpdate, AlineacionUpdate
from config import ALLOWED_EXCEL_MIME_TYPES
from security import get_current_user, can_edit_ports, can_upload_excel, registrar_auditoria

router = APIRouter(prefix="/api", tags=["Cabezales y Alineación"])

@router.get("/cabezales")
def get_cabezales(ciudad: str = None, id_equipo: str = None, db: Session = Depends(get_db)):
    query = db.query(CabezalModel)
    if ciudad: query = query.filter(CabezalModel.ciudad.ilike(f"%{ciudad}%"))
    if id_equipo: query = query.filter(CabezalModel.id_equipo.ilike(f"%{id_equipo}%"))
    return {"status": "success", "data": query.all()}

@router.get("/cabezales/{cabezal_id}/alineacion")
def get_alineacion(cabezal_id: int, db: Session = Depends(get_db)):
    return {"status": "success", "data": db.query(AlineacionCabezalModel).filter(AlineacionCabezalModel.cabezal_id == cabezal_id).order_by(AlineacionCabezalModel.id.asc()).all()}

@router.put("/cabezales/{cabezal_id}")
def update_cabezal(cabezal_id: int, data: CabezalUpdate, current_user: UserModel = Depends(get_current_user), db: Session = Depends(get_db)):
    if not can_edit_ports(current_user): raise HTTPException(status_code=403)
    cabezal = db.query(CabezalModel).filter(CabezalModel.id == cabezal_id).first()
    if not cabezal: raise HTTPException(status_code=404)
    for key, val in data.model_dump(exclude_unset=True).items(): setattr(cabezal, key, val)
    db.commit()
    return {"status": "success"}

@router.delete("/cabezales/{cabezal_id}")
def delete_cabezal(cabezal_id: int, current_user: UserModel = Depends(get_current_user), db: Session = Depends(get_db)):
    if not can_edit_ports(current_user): raise HTTPException(status_code=403)
    cabezal = db.query(CabezalModel).filter(CabezalModel.id == cabezal_id).first()
    if cabezal:
        db.delete(cabezal)
        db.commit()
    return {"status": "success"}

@router.put("/alineaciones/{alineacion_id}")
def update_alineacion(alineacion_id: int, data: AlineacionUpdate, current_user: UserModel = Depends(get_current_user), db: Session = Depends(get_db)):
    if not can_edit_ports(current_user): raise HTTPException(status_code=403)
    alineacion = db.query(AlineacionCabezalModel).filter(AlineacionCabezalModel.id == alineacion_id).first()
    if not alineacion: raise HTTPException(status_code=404)
    for key, val in data.model_dump(exclude_unset=True).items(): setattr(alineacion, key, val)
    db.commit()
    return {"status": "success"}

@router.delete("/alineaciones/{alineacion_id}")
def delete_alineacion(alineacion_id: int, current_user: UserModel = Depends(get_current_user), db: Session = Depends(get_db)):
    if not can_edit_ports(current_user): raise HTTPException(status_code=403)
    alineacion = db.query(AlineacionCabezalModel).filter(AlineacionCabezalModel.id == alineacion_id).first()
    if alineacion:
        db.delete(alineacion)
        db.commit()
    return {"status": "success"}

@router.post("/cabezales/upload-excel")
async def upload_cabezales_excel(
    file: UploadFile = File(...), 
    mode: str = Query("preview"), 
    ciudad: Optional[str] = Query(None), 
    current_user: UserModel = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    try:
        # 1. Validaciones de seguridad
        if not can_upload_excel(current_user): 
            return JSONResponse(status_code=403, content={"status": "error", "detail": "Permisos insuficientes."})
        
        filename = file.filename or ""
        content_type = file.content_type or ""
        
        if not (content_type in ALLOWED_EXCEL_MIME_TYPES or filename.lower().endswith(('.xlsx', '.xls'))): 
            return JSONResponse(status_code=400, content={"status": "error", "detail": "El archivo no es un Excel válido."})
        
        # 2. Lectura del archivo
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents)).fillna("")
        column_headers = [str(col).upper().strip() for col in df.columns]

        def get_idx(targets): return next((column_headers.index(t) for t in targets if t in column_headers), -1)

        # Buscar columnas principales
        idx_ciudad = get_idx(["CIUDAD"])
        idx_id = get_idx(["ID", "ID EQUIPO", "ID_EQUIPO", "CABEZAL ID"])
        idx_servicio = get_idx(["SERVICIO", "CLIENTE", "CLIENTE / SERVICIO"])
        
        # Buscar columnas de Hardware e IP
        idx_marca = get_idx(["MARCA", "MARCA/MOD"])
        idx_modelo = get_idx(["MODELO"])
        idx_serie = get_idx(["SERIE", "NO. SERIE", "NUMERO DE SERIE"])
        idx_ip = get_idx(["GESTION QAM", "IP", "IP GESTION", "GESTION"])
        
        # Buscar columnas de Alineación
        idx_portadora = get_idx(["PORTADORA", "PORTADORA / CANAL"])
        idx_canal = get_idx(["# CANAL", "CANAL NUM", "CANAL"])
        idx_nombre = get_idx(["NOMBRE DE CANAL", "NOMBRE CANAL"])

        if idx_id == -1 or idx_servicio == -1: 
            return JSONResponse(status_code=400, content={"status": "error", "detail": f"Columnas faltantes (Se requiere ID y Servicio). Encontradas: {column_headers}"})

        preview_data = []
        has_errors = False
        
        for _, row in df.iterrows():
            vals = list(row.values)
            def read_val(idx): return str(vals[idx]).strip() if idx != -1 and idx < len(vals) else ""
            
            val_id = read_val(idx_id)
            val_servicio = read_val(idx_servicio)
            val_ciudad = read_val(idx_ciudad) if idx_ciudad != -1 and read_val(idx_ciudad) else (ciudad or "Sin Ciudad") 
            
            # Datos de Hardware
            val_marca = read_val(idx_marca)
            val_modelo = read_val(idx_modelo)
            val_serie = read_val(idx_serie)
            val_ip = read_val(idx_ip)
            
            # Datos de Alineación
            val_portadora = read_val(idx_portadora)
            val_canal = read_val(idx_canal)
            val_nombre = read_val(idx_nombre)
            
            if not val_id and not val_servicio and not val_canal: continue
            
            errores_fila = []
            if not val_id: errores_fila.append("Falta ID de Equipo.")
            
            valido = len(errores_fila) == 0
            if not valido: has_errors = True
            
            preview_data.append({
                "ID_EQUIPO": val_id, 
                "CIUDAD": val_ciudad, 
                "SERVICIO": val_servicio,
                "MARCA": val_marca,
                "MODELO": val_modelo,
                "SERIE": val_serie,
                "GESTION_QAM": val_ip,
                "PORTADORA": val_portadora,
                "CANAL": val_canal, 
                "NOMBRE_CANAL": val_nombre, 
                "_errores": errores_fila, 
                "_valido": valido
            })

        # === 1. MODO PREVISUALIZACIÓN ===
        if mode == "preview": 
            return {"status": "success", "data": preview_data, "has_errors": has_errors}

        # === 2. MODO INYECCIÓN A BASE DE DATOS ===
        for row in preview_data:
            if not row["_valido"]: continue
            
            # A) GESTIÓN DEL CABEZAL
            cab_db = db.query(CabezalModel).filter(
                CabezalModel.ciudad == row["CIUDAD"], 
                CabezalModel.id_equipo == row["ID_EQUIPO"]
            ).first()
            
            if not cab_db:
                cab_db = CabezalModel(ciudad=row["CIUDAD"], id_equipo=row["ID_EQUIPO"])
                db.add(cab_db)
            
            # Asignación segura de propiedades (evita crasheos si la columna no existe en tu Models)
            cab_db.servicio = row["SERVICIO"]
            if row["MARCA"] and hasattr(cab_db, 'marca'): cab_db.marca = row["MARCA"]
            if row["MODELO"] and hasattr(cab_db, 'modelo'): cab_db.modelo = row["MODELO"]
            if row["SERIE"] and hasattr(cab_db, 'serie'): cab_db.serie = row["SERIE"]
            
            # Busca la propiedad de IP (diferentes nombres posibles en modelos)
            if row["GESTION_QAM"]:
                if hasattr(cab_db, 'gestion_qam'): cab_db.gestion_qam = row["GESTION_QAM"]
                elif hasattr(cab_db, 'ip_gestion'): cab_db.ip_gestion = row["GESTION_QAM"]
                elif hasattr(cab_db, 'ip'): cab_db.ip = row["GESTION_QAM"]

            # Forzar guardado para obtener el ID real del cabezal (necesario para la alineación)
            db.flush() 
            
            # B) GESTIÓN DE LA ALINEACIÓN
            if row["CANAL"] or row["PORTADORA"] or row["NOMBRE_CANAL"]:
                al_db = db.query(AlineacionCabezalModel).filter(
                    AlineacionCabezalModel.cabezal_id == cab_db.id,
                    AlineacionCabezalModel.canal_num == row["CANAL"]
                ).first()
                
                if not al_db:
                    al_db = AlineacionCabezalModel(cabezal_id=cab_db.id, canal_num=row["CANAL"])
                    db.add(al_db)
                
                if row["PORTADORA"] and hasattr(al_db, 'portadora'): al_db.portadora = row["PORTADORA"]
                if row["NOMBRE_CANAL"] and hasattr(al_db, 'nombre_canal'): al_db.nombre_canal = row["NOMBRE_CANAL"]
                
        db.commit()
        
        ciudad_audit = ciudad if ciudad else "Múltiples Ciudades"
        registrar_auditoria(db, current_user.username, "CARGA CABEZALES", "CABEZALES", f"Excel procesado en {ciudad_audit}.")
        
        return {"status": "success", "detail": f"Proceso completado. Se inyectaron {len(preview_data)} registros de equipos y alineación."}
        
    except Exception as e:
        db.rollback()
        error_details = traceback.format_exc()
        print(error_details)
        return JSONResponse(status_code=500, content={"status": "error", "detail": f"Error: {str(e)}\n{error_details}"})

        # === 1. MODO PREVISUALIZACIÓN ===
        if mode == "preview": 
            return {"status": "success", "data": preview_data, "has_errors": has_errors}

        # === 2. MODO INYECCIÓN A BASE DE DATOS ===
        for row in preview_data:
            if not row["_valido"]: continue
            
            cab_db = db.query(CabezalModel).filter(
                CabezalModel.ciudad == row["CIUDAD"], 
                CabezalModel.id_equipo == row["ID_EQUIPO"]
            ).first()
            
            if cab_db:
                # Si existe, lo actualiza
                cab_db.servicio = row["SERVICIO"]
            else:
                # Si no existe, lo crea
                nuevo_cab = CabezalModel(
                    ciudad=row["CIUDAD"],
                    id_equipo=row["ID_EQUIPO"],
                    servicio=row["SERVICIO"]
                )
                db.add(nuevo_cab)
                
        db.commit()
        
        # Auditoría con manejo por si no hay ciudad
        ciudad_audit = ciudad if ciudad else "Múltiples Ciudades (desde archivo)"
        registrar_auditoria(db, current_user.username, "CARGA CABEZALES", "CABEZALES", f"Se actualizaron cabezales en {ciudad_audit} mediante Excel.")
        
        return {"status": "success", "detail": f"Proceso completado. Se inyectaron/actualizaron {len(preview_data)} equipos."}
        
    except Exception as e:
        db.rollback()
        # ESTO ES LA MAGIA: Captura el error de Python exacto y lo envía de vuelta
        error_details = traceback.format_exc()
        print(error_details)
        return JSONResponse(status_code=500, content={"status": "error", "detail": f"Excepción de Python: {str(e)} \n\n {error_details}"})

# ================= EXPORTACIÓN EXCEL ALINEACIÓN =================
@router.get("/cabezales/{cabezal_id}/exportar-excel")
def exportar_alineacion_excel(cabezal_id: int, current_user: UserModel = Depends(get_current_user), db: Session = Depends(get_db)):
    try:
        cabezal = db.query(CabezalModel).filter(CabezalModel.id == cabezal_id).first()
        if not cabezal: raise HTTPException(status_code=404)
        
        alineaciones = db.query(AlineacionCabezalModel).filter(AlineacionCabezalModel.cabezal_id == cabezal_id).order_by(AlineacionCabezalModel.id.asc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Alineacion_Canales"
        headers = ["PORTADORA", "FORMATO", "CANAL NUM", "NOMBRE DE CANAL", "MCAST IP", "SOURCE IP", "UDP", "SID"]
        ws.append(headers)
        
        for col_idx in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0B132B")
            
        for al in alineaciones: 
            ws.append([al.portadora, al.formato, al.canal_num, al.nombre_canal, al.mcast_ip, al.source_ip, al.udp, al.sid])
            
        for col in ws.columns: 
            ws.column_dimensions[col[0].column_letter].width = 18
            
        if len(alineaciones) > 0:
            tab = Table(displayName="TablaAlineacion", ref=f"A1:H{len(alineaciones)+1}")
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(tab)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            headers={
                "Content-Disposition": f"attachment; filename=Alineacion_{cabezal.servicio}.xlsx", 
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except Exception as e: 
        return JSONResponse(status_code=500, content={"status": "error", "detail": str(e)})